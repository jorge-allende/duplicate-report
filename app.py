from flask import Flask, request, send_file, render_template, jsonify
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
import logging

# Initialize Flask app
app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
PROCESSED_FOLDER = 'processed'
PROCESSED_FILE = 'processed_file.xlsx'

# Set up logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Ensure folders exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)

def process_excel(file_path, output_path):
    try:
        # Load the file
        if file_path.endswith('.csv'):
            logger.info("Detected CSV file. Reading as CSV.")
            df = pd.read_csv(file_path)
        elif file_path.endswith(('.xls', '.xlsx')):
            logger.info("Detected Excel file. Reading as Excel.")
            df = pd.read_excel(file_path, engine='openpyxl')
        else:
            logger.error("Unsupported file format: %s", file_path)
            raise ValueError("Unsupported file format. Only .csv, .xls, and .xlsx files are supported.")

        # Step 1: Remove specified columns
        columns_to_remove = [
            'Agreement Pending or Approved?',
            'Last Updated Date',
            'Salesforce Object ID',
            'Within A P/C Hierarchy',
            'Analyze Ingestion Source',
            'File Type',
            'Attachment Count'
        ]
        df.drop(columns=[col for col in columns_to_remove if col in df.columns], inplace=True)

        # Save to a temporary file for openpyxl processing
        temp_path = "temp.xlsx"
        df.to_excel(temp_path, index=False)

        # Load workbook with openpyxl
        wb = load_workbook(temp_path)
        ws = wb.active

        # Rename columns as per final report
        ws['A1'].value = 'Agreement UUID'
        ws['B1'].value = 'Link'
        ws['C1'].value = 'Creation Date'
        ws['D1'].value = 'Agreement Name'

        # Step 2-7: Apply Conditional Formatting to highlight duplicates in Column F
        col_idx = 6  # Column F (Content Checksum)
        col_letter = get_column_letter(col_idx)
        formula = f"=COUNTIF(${col_letter}$2:{col_letter}2, {col_letter}2)>1"
        rose_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
        dxf = DifferentialStyle(fill=rose_fill)
        rule = Rule(type="expression", dxf=dxf, formula=[formula])
        ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)

        # Step 14: Hide column E (Original File Name)
        ws.column_dimensions['E'].hidden = True

        # Step 15: Add 'Delete? (X)' in Column G and format it like other headers
        ws['G1'] = 'Delete? (X)'
        ws['G1'].font = ws['A1'].font.copy(bold=True)  # Match bold font style
        ws['G1'].alignment = ws['A1'].alignment.copy()  # Match alignment

        # Add black border to all headers (A1 to G1)
        black_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        for col in range(1, 8):  # Columns A to G
            cell = ws.cell(row=1, column=col)
            cell.border = black_border

        # Save the workbook
        wb.save(output_path)
        logger.info("Processing complete. Output saved to: %s", output_path)

    except Exception as e:
        logger.exception("Error processing the file")

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        logger.warning("No file part in the request")
        return 'No file part', 400
    file = request.files['file']
    if file.filename == '':
        logger.warning("No file selected for upload")
        return 'No selected file', 400
    if file:
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)
        logger.info("File uploaded: %s", filepath)
        # Process the file
        processed_path = os.path.join(PROCESSED_FOLDER, PROCESSED_FILE)
        process_excel(filepath, processed_path)
        return jsonify(status='File uploaded and processed'), 200

@app.route('/download')
def download_file():
    processed_file_path = os.path.join(PROCESSED_FOLDER, PROCESSED_FILE)
    logger.info("Downloading file from: %s", processed_file_path)
    if not os.path.exists(processed_file_path):
        logger.error("Processed file not found: %s", processed_file_path)
        return 'File not found', 404
    return send_file(processed_file_path, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
